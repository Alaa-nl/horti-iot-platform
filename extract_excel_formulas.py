#!/usr/bin/env python3
"""
Extract formulas from Excel files for analysis
"""

import openpyxl
import json
import sys
from pathlib import Path

def extract_excel_formulas(excel_path):
    """Extract all formulas and values from an Excel file"""

    try:
        # Load the workbook
        wb = openpyxl.load_workbook(excel_path, data_only=False)

        result = {}

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            result[sheet_name] = {
                'formulas': {},
                'values': {},
                'constants': {}
            }

            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cell_ref = f"{cell.column_letter}{cell.row}"

                        # Check if it's a formula
                        if isinstance(cell.value, str) and cell.value.startswith('='):
                            result[sheet_name]['formulas'][cell_ref] = cell.value
                        else:
                            # It's a constant or calculated value
                            result[sheet_name]['values'][cell_ref] = str(cell.value)

                            # Try to get the actual value if formula
                            try:
                                wb_values = openpyxl.load_workbook(excel_path, data_only=True)
                                sheet_values = wb_values[sheet_name]
                                cell_with_value = sheet_values[cell_ref]
                                if cell_with_value.value != cell.value:
                                    result[sheet_name]['formulas'][cell_ref] = cell.value
                                    result[sheet_name]['values'][cell_ref] = str(cell_with_value.value)
                            except:
                                pass

        return result

    except Exception as e:
        return {'error': str(e)}

def save_formulas_to_text(excel_path, output_path=None):
    """Save extracted formulas to a readable text file"""

    if output_path is None:
        output_path = Path(excel_path).with_suffix('.formulas.txt')

    formulas = extract_excel_formulas(excel_path)

    with open(output_path, 'w') as f:
        for sheet_name, sheet_data in formulas.items():
            if sheet_name == 'error':
                f.write(f"Error: {sheet_data}\n")
                continue

            f.write(f"\n{'='*60}\n")
            f.write(f"SHEET: {sheet_name}\n")
            f.write(f"{'='*60}\n\n")

            if sheet_data['formulas']:
                f.write("FORMULAS:\n")
                f.write("---------\n")
                for cell, formula in sorted(sheet_data['formulas'].items()):
                    f.write(f"{cell}: {formula}\n")
                f.write("\n")

            # Only write key constants (numeric values that might be parameters)
            if sheet_data['values']:
                f.write("KEY VALUES/CONSTANTS:\n")
                f.write("--------------------\n")
                for cell, value in sorted(sheet_data['values'].items()):
                    try:
                        # Only include numeric values that might be constants
                        float_val = float(value)
                        if 0.001 <= abs(float_val) <= 10000:  # Reasonable range for constants
                            f.write(f"{cell}: {value}\n")
                    except:
                        # Include labels in first rows/columns
                        col = cell[0]
                        row = int(cell[1:])
                        if row <= 5 or col == 'A':
                            f.write(f"{cell}: {value}\n")

    print(f"Formulas extracted to: {output_path}")
    return output_path

if __name__ == "__main__":
    # Process all Excel files in Roel Sheets directory
    sheets_dir = Path("/Users/alaadrobe/Desktop/Intrenship/Roel Sheets")

    if sheets_dir.exists():
        for excel_file in sheets_dir.glob("*.xlsx"):
            print(f"\nProcessing: {excel_file.name}")
            output = save_formulas_to_text(excel_file)
            print(f"  Saved to: {output}")
    else:
        print(f"Directory not found: {sheets_dir}")
        print("\nUsage: python3 extract_excel_formulas.py")
        print("Or provide path: python3 extract_excel_formulas.py /path/to/file.xlsx")

        if len(sys.argv) > 1:
            excel_path = sys.argv[1]
            if Path(excel_path).exists():
                save_formulas_to_text(excel_path)